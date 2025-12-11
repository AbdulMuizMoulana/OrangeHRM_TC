import openpyxl


def get_rows_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_row


def get_columns_index_by_name(file, sheet_name, column_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]

    for col in range(1, sheet.max_column + 1):
        if sheet.cell(1, col).value == column_name:
            return col

    raise Exception(f"column {column_name} Not found in the header")


def read_data(file, sheet_name, row_num, column_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.cell(row_num, column_name).value


def write_data(file, sheet_name, row_num, column_name, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    sheet.cell(row_num, column_name).value = data
    workbook.save(file)
